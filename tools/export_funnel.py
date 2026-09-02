#!/usr/bin/env python3
"""שני קובצי אקסל: משפך המיון המלא, והפער בין ערוצי הגיוס.

הכול נקרא מ-data/recruitment_data.json - כלומר מאותם מספרים בדיוק
שהמחשבון מציג. אין כאן חישוב נוסף ואין מספרים שנכתבו ידנית.

  משפך מיון מלא.xlsx   - שלב אחר שלב: כמה מועמדים ייחודיים נמדדו בכל
                         שלב, איזה חלק עובר אליו מהשלב שלפניו, ואיזה
                         חלק עובר אליו מההגשה. שתי המדידות ישירות.
  פער בין ערוצי גיוס.xlsx - אותו משפך בנפרד לכל ערוץ, ולצידו הפער.

הרצה:  python3 tools/export_funnel.py
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT / "data" / "recruitment_data.json"
OUT_DIR = ROOT / "קבצים" / "התקבל"

FUNNEL_OUT = OUT_DIR / "משפך מיון מלא.xlsx"
SEGMENTS_OUT = OUT_DIR / "פער בין ערוצי גיוס.xlsx"

THIN = Side(style="thin", color="D0D7E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HIRE_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE_FILL = PatternFill("solid", fgColor="F7F9FC")


def sheet(wb, title, head, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.rightToLeft = True
    for c, (name, width) in enumerate(head, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 42
    return ws


def put(ws, row, vals, fmt=None, fill=None):
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if fmt and c in fmt:
            cell.number_format = fmt[c]
        if isinstance(v, str) and len(v) > 30:
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def note(ws, row, width, text, height_rows=6):
    cell = ws.cell(row=row, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = NOTE_FILL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row + height_rows, end_column=width)


def reach(m):
    return None if m is None else m["reach"]["mid"]


def med(m):
    return None if m is None else m["days"]["median"]


def measured(m):
    return None if m is None else m["days"]["n"]


# ----------------------------------------------------------------------
# משפך המיון המלא
# ----------------------------------------------------------------------

FUNNEL_HEAD = [
    ("שלב", 22), ("מועמדים ייחודיים", 15), ("מהתאריך", 12), ("עד התאריך", 12),
    ("מהשלב הקודם", 20), ("עוברים מהקודם", 14), ("חציון ימים מהקודם", 14),
    ("נמדד על", 11),
    ("עוברים מההגשה", 14), ("חציון ימים מההגשה", 15), ("נמדד על ", 11),
]

FUNNEL_FMT = {2: "#,##0", 6: "0.0%", 7: "0.0", 8: "#,##0",
              9: "0.0%", 10: "0.0", 11: "#,##0"}


def funnel_rows(ws, funnel, start=2):
    r = start
    for row in funnel["rows"]:
        put(ws, r, [
            row["label"], row["candidates"], row["first_date"], row["last_date"],
            "—" if row["from_prev"] is None else row["from_prev"]["label"],
            reach(row["from_prev"]), med(row["from_prev"]),
            measured(row["from_prev"]),
            reach(row["from_first"]), med(row["from_first"]),
            measured(row["from_first"]),
        ], FUNNEL_FMT, HIRE_FILL if row["is_hire"] else None)
        r += 1
    return r


def build_funnel_book(data):
    wb = Workbook()
    f = data["funnel"]
    first_label = f["rows"][0]["label"]

    ws = sheet(wb, "משפך המיון", FUNNEL_HEAD, first=True)
    ws.freeze_panes = "B2"
    r = funnel_rows(ws, f)

    note(ws, r + 1, len(FUNNEL_HEAD), (
        "איך לקרוא את הטבלה\n"
        "«מועמדים ייחודיים» הוא כמה מועמדים שונים הופיעו אי פעם בשלב הזה "
        "בקבצים - היקף התנועה בפועל, ולא קוהורט מדוד. המספר אינו יורד "
        "בהכרח משלב לשלב, כי המשפך אינו סדרתי: יש מועמדים שנכנסים באמצע "
        "התהליך ויש שמדלגים על שלב.\n\n"
        "«עוברים מהקודם» ו«עוברים מההגשה» הם שתי מדידות נפרדות, ושתיהן "
        "ישירות: איזה חלק מהמועמדים שהיו בשלב האחד הגיעו בפועל לשלב השני "
        "אחר כך. «עוברים מההגשה» אינו מכפלה של «עוברים מהקודם» - שרשור "
        "אחוזים היה מייצר מספר שאינו קיים בנתונים.\n\n"
        "כל שיעור נמדד בשני בסיסים - חלון זמן קבוע, וקוהורט שהספיק "
        "להבשיל - והערך שכאן הוא הממוצע ביניהם. הסיבה: הנתונים קטועים "
        "מימין, ומי שביצע פעילות סמוך לסוף התקופה עדיין לא הספיק להתקדם.\n\n"
        "«חציון ימים» נמדד רק על מי שהגיע בפועל לשלב היעד. מי שנעצר בדרך "
        "אינו נספר. «נמדד על» הוא מספר המועמדים שעליהם נמדד הזמן הזה.\n\n"
        "השלב הראשון הוא «" + first_label + "», ולכן אין לו שלב קודם."), 10)

    meta_ws = sheet(wb, "מקור הנתונים", [("פריט", 34), ("ערך", 60)], first=False)
    m = data["meta"]
    rows = [
        ("נבנה בתאריך", m["generated_at"]),
        ("קובצי פעילות", " · ".join(Path(x).name for x in m["activities_files"])),
        ("שורות פעילות", m["activity_rows"]),
        ("שורות כפולות שנופו", m["activity_duplicate_rows_dropped"]),
        ("מועמדים ייחודיים בפעילות", m["activity_candidates"]),
        ("תקופת הפעילות", m["activity_first"] + " עד " + m["activity_last"]),
        ("קובץ הגיוסים", Path(m["hires_file"]).name),
        ("שורות גיוס", m["hire_rows"]),
        ("מגויסים ייחודיים", m["hire_candidates"]),
        ("תקופת הגיוסים", m["hire_first"] + " עד " + m["hire_last"]),
        ("מגויסים שאינם בקובץ הפעילויות", m["hires_without_activity"]),
        ("סוגי פעילות שלא מופו לשלב",
         ", ".join(m["unmapped_activity_types"]) or "אין"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        put(meta_ws, i, [k, v], {2: "#,##0"} if isinstance(v, int) else None)

    return wb


# ----------------------------------------------------------------------
# הפער בין ערוצי הגיוס
# ----------------------------------------------------------------------

GAP_HEAD = [
    ("ערוץ", 34), ("מגישים", 12), ("מגויסים", 12),
    ("מהגשה עד גיוס", 14), ("חציון ימים", 12), ("נמדד על", 11),
    ("פי כמה מכלל המועמדים", 16),
]

GAP_FMT = {2: "#,##0", 3: "#,##0", 4: "0.00%", 5: "0.0", 6: "#,##0", 7: "0.00"}


def build_segments_book(data):
    wb = Workbook()
    segs = [g for g in data["segments"] if g["funnel"]]
    base = next((g for g in segs if g["key"] == "all"), None)
    base_rate = None if base is None else reach(base["funnel"]["rows"][-1]["from_first"])

    ws = sheet(wb, "הפער בין הערוצים", GAP_HEAD, first=True)
    r = 2
    for g in segs:
        rows = g["funnel"]["rows"]
        hire = rows[-1]
        ff = hire["from_first"]
        rate = reach(ff)
        put(ws, r, [
            g["label"], rows[0]["candidates"], hire["candidates"],
            rate, med(ff), measured(ff),
            None if (rate is None or not base_rate) else round(rate / base_rate, 2),
        ], GAP_FMT, HIRE_FILL if g["key"] == "all" else None)
        r += 1

    note(ws, r + 1, len(GAP_HEAD), (
        "מה הטבלה משווה\n"
        "«מגישים» הוא מספר המועמדים הייחודיים שנרשמה להם הגשה בערוץ הזה, "
        "ו«מגויסים» הוא כמה מהם התגייסו בפועל. «מהגשה עד גיוס» הוא שיעור "
        "ההגעה מההגשה אל הגיוס, נמדד ישירות בין שני השלבים האלה.\n\n"
        "כל מועמד משויך לערוץ אחד בלבד, לפי עמודת «דרישה» שברשומת ההגשה "
        "שלו. למי שאין רשומת הגשה בקבצים נלקחת הדרישה מהפעילות המוקדמת "
        "ביותר שלו. לכן הערוצים אינם חופפים, וסכומם קטן מכלל המועמדים - "
        "יש ערוצים נוספים שאינם בטבלה הזו.\n\n"
        "«פי כמה מכלל המועמדים» הוא היחס בין שיעור הגיוס של הערוץ לשיעור "
        "הגיוס של כלל המועמדים. זה המספר שאומר איזה ערוץ מביא מועמדים "
        "שמתגייסים, ולא רק מועמדים.\n\n"
        "השוואה זו אינה טענה על איכות המועמד. ערוץ יכול להיות טוב יותר "
        "גם מפני שהוא מגיע לאוכלוסייה מתאימה יותר, וגם מפני שהמועמדים "
        "שלו מטופלים אחרת בתהליך. הנתונים אינם מפרידים בין השניים."), 10)

    for g in segs:
        # שם גיליון באקסל אינו יכול להכיל / \ ? * [ ] : ולא לעבור 31 תווים
        title = "".join(" " if ch in "/\\?*[]:" else ch for ch in g["label"])
        ws = sheet(wb, title.strip()[:28] or g["key"], FUNNEL_HEAD)
        ws.freeze_panes = "B2"
        end = funnel_rows(ws, g["funnel"])
        note(ws, end + 1, len(FUNNEL_HEAD),
             "משפך המיון המלא של «" + g["label"] + "» בלבד. "
             + ("כל המועמדים, בלי סינון לפי דרישה."
                if g["match"] is None
                else "כולל את הדרישות: " + ", ".join(g["match"]) + ".")
             + " שאר ההסברים זהים לגיליון «משפך המיון» שבקובץ משפך המיון המלא.",
             3)

    return wb


def main():
    if not DATA_PATH.exists():
        sys.exit(f"מאגר הנתונים חסר: {DATA_PATH}. יש להריץ make.")
    with DATA_PATH.open(encoding="utf-8") as fh:
        data = json.load(fh)

    if not data.get("funnel"):
        sys.exit("אין נתוני משפך במאגר. יש להריץ make.")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    build_funnel_book(data).save(FUNNEL_OUT)
    print(f"נכתב {FUNNEL_OUT.relative_to(ROOT)} "
          f"({len(data['funnel']['rows'])} שלבים)")

    if data.get("segments"):
        build_segments_book(data).save(SEGMENTS_OUT)
        print(f"נכתב {SEGMENTS_OUT.relative_to(ROOT)} "
              f"({len([g for g in data['segments'] if g['funnel']])} ערוצים)")
    else:
        print("אין פילוח ערוצים במאגר - הקובץ השני לא נוצר.")


if __name__ == "__main__":
    main()
