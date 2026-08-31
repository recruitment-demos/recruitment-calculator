#!/usr/bin/env python3
"""דוח זמנים בין כל שני שלבים, כולל סטיית תקן וחלון הסינון.

מייצר אקסל עם שורה לכל מעבר: מכל שלב לכל שלב שאחריו, ומכל שלב עד
הגיוס. לכל שורה - ממוצע, סטיית תקן, וגבולות החלון.

מוצגים שני חלונות, כי הם עונים על שאלות שונות:

  חלון רגיל   ממוצע ± 1.5σ על הימים עצמם. זה מה שנתבקש במפורש.
  חלון לוג    ממוצע ± 1.5σ על log(ימים+1), כלומר חלון כפלי.

הסיבה לשני החלונות: התפלגות זמנים היא א-סימטרית - יש גבול תחתון של
אפס ואין גבול עליון. בהתפלגות כזו ממוצע פחות 1.5σ יוצא לעתים קרובות
שלילי, ואז החלון אינו חוסם כלום בקצה המהיר. החלון על הלוג הוא הצורה
הנכונה לחסום התפלגות כזו משני הצדדים.

הרצה:  python3 tools/analyze_durations.py
"""

import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "tools"))

from build_dataset import days_between, load_config, load_sources  # noqa: E402

OUT = ROOT / "קבצים" / "התקבל" / "זמנים בין שלבים.xlsx"
SIGMAS = 1.5


def window_stats(days, sigmas=SIGMAS):
    """סטטיסטיקה וחלון, בשני קני מידה."""
    d = np.asarray(days, dtype=float)
    n = len(d)
    mean = float(d.mean())
    sd = float(d.std(ddof=1)) if n > 1 else 0.0
    lo, hi = mean - sigmas * sd, mean + sigmas * sd

    lg = np.log(d + 1.0)
    lmean, lsd = float(lg.mean()), float(lg.std(ddof=1)) if n > 1 else 0.0
    llo = float(np.exp(lmean - sigmas * lsd) - 1.0)
    lhi = float(np.exp(lmean + sigmas * lsd) - 1.0)

    return {
        "n": n,
        "mean": mean,
        "sd": sd,
        "median": float(np.median(d)),
        "min": float(d.min()),
        "max": float(d.max()),
        "lo": lo,
        "hi": hi,
        "inside": float(((d >= lo) & (d <= hi)).mean()),
        "log_lo": llo,
        "log_hi": lhi,
        "log_inside": float(((d >= llo) & (d <= lhi)).mean()),
        "cut_fast": float((d < llo).mean()),
        "cut_slow": float((d > lhi).mean()),
        "within_2": float((d <= 2).mean()),
        "within_7": float((d <= 7).mean()),
    }


def collect():
    cfg = load_config()
    act, rec = load_sources(cfg)
    hire_date = rec.groupby("candidate")["date"].min()

    first = {}
    for s in cfg["stages"]:
        if s["activity_type"] is None:
            continue
        rows = act[act["activity"] == s["activity_type"]]
        first[s["key"]] = rows.groupby("candidate")["date"].min()

    keys = [s["key"] for s in cfg["stages"] if s["activity_type"] is not None]
    label = {s["key"]: s["label"] for s in cfg["stages"]}
    label["hire"] = "גיוס"

    rows = []
    # מכל שלב עד הגיוס - השורות שהמחשבון נשען עליהן
    for k in keys:
        d = days_between(first[k], hire_date)
        if len(d) < 2:
            continue
        rows.append(("עד גיוס", label[k], "גיוס", window_stats(d)))
    # בין כל שני שלבים
    for i, a in enumerate(keys):
        for b in keys[i + 1:]:
            d = days_between(first[a], first[b])
            if len(d) < 2:
                continue
            rows.append(("בין שלבים", label[a], label[b], window_stats(d)))
    return rows


HEAD = [
    ("סוג", 12), ("משלב", 20), ("אל שלב", 20), ("תצפיות", 10),
    ("ממוצע ימים", 12), ("סטיית תקן", 11), ("חציון", 9),
    ("מינימום 1.5σ", 13), ("מקסימום 1.5σ", 14), ("% בתוך החלון", 13),
    ("מינימום לוג 1.5σ", 16), ("מקסימום לוג 1.5σ", 17), ("% בתוך חלון הלוג", 16),
    ("% נחתך כמהיר מדי", 17), ("% נחתך כאיטי מדי", 17),
    ("% תוך יומיים", 13), ("% תוך שבוע", 12),
    ("המהיר ביותר", 12), ("האיטי ביותר", 12),
]


IMPACT_HEAD = [
    ("שלב", 20), ("רצפה שנמצאה (ימים)", 17),
    ("רצפת 1.5σ להשוואה", 17), ("תצפיות לפני", 12),
    ("נפסלו", 10), ("תצפיות אחרי", 12),
    ("שיעור גיוס", 12), ("חציון ימים", 12), ("ממוצע ימים", 12),
    ("הנימוק", 60),
]


def impact_sheet(wb):
    """מה הסינון עשה בפועל, לפי מאגר הנתונים הבנוי."""
    path = ROOT / "data" / "recruitment_data.json"
    if not path.exists():
        return
    import json
    data = json.loads(path.read_text(encoding="utf-8"))
    f = data.get("duration_filter", {})

    ws = wb.create_sheet("השפעת הסינון")
    ws.sheet_view.rightToLeft = True
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    head = ws.cell(row=1, column=1, value=(
        f"השיטה: {f.get('method')}. הרצפה נמצאת מצורת ההתפלגות - בליטה "
        f"בימים הראשונים, שקע אחריה, ואז האוכלוסייה האמיתית - ונקבעת בשקע. "
        f"תא בהיסטוגרמה {f.get('bin_days')} ימים, יחס בליטה לשקע "
        f"{f.get('spike_ratio')}. שלב שההתפלגות שלו חלקה אינו נחתך כלל. "
        "תצפית שנפסלה אינה נספרת לא בזמנים ולא בשיעור ההגעה."))
    head.font = Font(bold=True, size=11)
    head.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=2,
                   end_column=len(IMPACT_HEAD))

    for c, (name, width) in enumerate(IMPACT_HEAD, start=1):
        cell = ws.cell(row=4, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[4].height = 40

    r = 5
    for st in data["stages"]:
        if not st["has_data"]:
            continue
        w = st["hire_window"]
        vals = [
            st["label"],
            w["from_days"] if w["from_days"] is not None else "לא נמצאה",
            w["sigma_floor_days"],
            w["observations_before"],
            w["dropped_fast"],
            w["observations_after"],
            st["hire_rate"]["mid"],
            st["days_to_hire"]["median"],
            st["days_to_hire"]["mean"],
            w["reason"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c == 7:
                cell.number_format = "0.00%"
            elif c in (2, 3, 8, 9):
                cell.number_format = "0.0"
            elif c in (4, 5, 6):
                cell.number_format = "#,##0"
            if c == 10:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1

    note = ws.cell(row=r + 2, column=1, value=(
        "למה סטיית תקן נפסלה כשיטת סינון: בבדיקת קבצים היא נותנת רצפה של "
        "4.6 ימים בלבד, ומשאירה גיוסים תוך שבוע מבדיקת קבצים - שאינם "
        "אפשריים בתהליך הזה. הסיבה היא שהתפלגות זמנים חסומה באפס ומשתרעת "
        "ימינה, ולכן «ממוצע פחות 1.5 סטיות תקן» יוצא שלילי או נמוך מאוד "
        "דווקא במעברים הארוכים. עמודת «רצפת 1.5σ להשוואה» מציגה מה היא "
        "היתה נותנת, לצד הרצפה שנמצאה בפועל.\n\n"
        "למה נחתך רק הקצה המהיר: בקצה האיטי אין בליטה מלאכותית. חיתוך שם "
        "היה מוחק גיוסים איטיים אמיתיים ומושך את הזמן הממוצע כלפי מטה - "
        "כלומר גורם למחשבון להבטיח גיוס מהיר מהמציאות.\n\n"
        "שלב שכתוב בו «לא נמצאה» הוא שלב שההתפלגות שלו חלקה, בלי שתי "
        "אוכלוסיות, ולכן לא נחתך בו דבר. כך נשמר למשל המעבר מיחב\"מ אל "
        "הגיוס, שבו גיוס תוך יומיים סביר לגמרי."))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 2, start_column=1,
                   end_row=r + 9, end_column=len(IMPACT_HEAD))


def main():
    rows = collect()
    wb = Workbook()
    ws = wb.active
    ws.title = "זמנים בין שלבים"
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")

    for c, (name, width) in enumerate(HEAD, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "D2"

    for r, (kind, a, b, s) in enumerate(rows, start=2):
        vals = [
            kind, a, b, s["n"],
            round(s["mean"], 1), round(s["sd"], 1), round(s["median"], 1),
            round(s["lo"], 1), round(s["hi"], 1), s["inside"],
            round(s["log_lo"], 1), round(s["log_hi"], 1), s["log_inside"],
            s["cut_fast"], s["cut_slow"], s["within_2"], s["within_7"],
            round(s["min"], 0), round(s["max"], 0),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c in (10, 13, 14, 15, 16, 17):
                cell.number_format = "0.0%"
            elif c >= 4:
                cell.number_format = "0.0"
            if c == 4:
                cell.number_format = "#,##0"
        # חלון שגבולו התחתון שלילי אינו חוסם כלום בקצה המהיר
        if s["lo"] < 0:
            ws.cell(row=r, column=8).fill = warn_fill

    note = ws.cell(row=len(rows) + 3, column=1, value=(
        "«מינימום 1.5σ» מסומן בצהוב כשהוא יוצא שלילי. במקרים האלה החלון "
        "הרגיל אינו חוסם דבר בקצה המהיר, כי אין ימים שליליים. לכן מוצג גם "
        "חלון על הלוג, שהוא הצורה הנכונה לחסום התפלגות זמנים משני הצדדים."))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=len(rows) + 3, start_column=1,
                   end_row=len(rows) + 5, end_column=len(HEAD))

    impact_sheet(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"נכתב {OUT} ({len(rows)} שורות)")


if __name__ == "__main__":
    main()
